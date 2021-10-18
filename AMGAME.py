from inspect import FullArgSpec
import random
from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty
from openpyxl import workbook, load_workbook
from kivy.uix.screenmanager import ScreenManager, Screen, ScreenManagerException
from kivy.lang import Builder
from kivy.uix.textinput import TextInput
import time

class am_grid(Screen, Widget):
    greenbelt = ObjectProperty(None)
    bluebelt = ObjectProperty(None)
    redbelt = ObjectProperty(None)
    yellowbelt = ObjectProperty(None)
    blowby = ObjectProperty(None)
    other = ObjectProperty(None)
    choicelabel = ObjectProperty(None)
    scorelable = ObjectProperty(None)
    message = ObjectProperty(None)
    lifelabel = ObjectProperty(None)
    namevar = ObjectProperty(None)
    
    wb = load_workbook('MASTEREXCEL.xlsx')
    wsAM = wb["AMZIPCODES"]
    score = 0
    lifes = 3
    now = time.strftime("%x")

        #this work for getting the values from the excel sheet and removes the none values 
    GreenBelt = []
    for col in wsAM['A']:
        GreenBelt.append(col.value)
    GreenBelt = [i for i in GreenBelt if i != None]

    BlueBelt = []
    for col in wsAM['B']:
        BlueBelt.append(col.value)
    blueBelt = [i for i in BlueBelt if i != None]

    RedBelt = []
    for col in wsAM['C']:
        RedBelt.append(col.value)
    RedBelt = [i for i in RedBelt if i != None]

    YellowBelt = []
    for col in wsAM['D']:
        YellowBelt.append(col.value)
    YellowBelt = [i for i in YellowBelt if i != None]

    BlowBy = []
    for col in wsAM['E']:
        BlowBy.append(col.value)
    BlowBy = [i for i in BlowBy if i != None]

    Other = []
    for col in wsAM['F']:
        Other.append(col.value)
    Other = [i for i in Other if i != None]

    fullList = list(GreenBelt + BlueBelt + RedBelt + YellowBelt + BlowBy + Other)
    choicevar = random.choice(fullList)   
    
    '''
    def check_answer(self):
        sortcode = self.choicevar
        correctlist = ""
        for i in range(self.fullList):
            if sortcode in i:
                correctlist = i
            
     '''       
            
    
    
    def GREENBELTCHECK (self,):
        answer = self.choicevar
        if answer in self.GreenBelt:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                wb = load_workbook('MASTEREXCEL.xlsx')
                ws = wb['USERS']
                ws.append({'A':self.namevar.text, 'B': self.score, 'C': self.now})
                wb.save("MASTEREXCEL.xlsx")
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def BLUEBELTCHECK (self,):
        answer = self.choicevar
        if answer in self.BlueBelt:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def REDBELTCHECK (self,):
        answer = self.choicevar
        if answer in self.RedBelt:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def YELLOWBELTCHECK (self,):
        answer = self.choicevar
        if answer in self.YellowBelt:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def BLOWBYCHECK (self,):
        answer = self.choicevar
        if answer in self.BlowBy:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def OTHERCHECK (self,):
        answer = self.choicevar
        if answer in self.Other:
            self.fullList.remove(answer)
            self.message.text = "correct"
            if self.fullList == []:
                fullList = list(self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other)
            self.choicevar = random.choice(self.fullList)
            self.choicelabel.text = str(self.choicevar)
            self.score += 1
            self.scorelabel.text = str(self.score)
        else:
            self.message.text = "wrong"
            self.score -= 1
            self.lifes -= 1
            if self.lifes == 0:
                App.get_running_app().stop()
            self.lifelabel.text =  "You have " + str(self.lifes) + " left"
            if self.score < 0:
                self.score = 0
            self.scorelabel.text = str(self.score)
    def reset(self,):
        self.score = 0
        self.lifes = 0
        self.choicevar = ""
        self.lifelabel.text =  "You have " + str(self.lifes) + " left"
        self.scorelabel.text = str(self.score)
        self.choicelabel.text = str(self.choicevar)